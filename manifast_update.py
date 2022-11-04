#!/usr/bin/env python
import os
import sys
import openpyxl
import xml.etree.ElementTree as ET
from CommentedTreeBuilder import CommentedTreeBuilder
import time

class Manifest_update:
	def __init__(self, version):
		self.file_name = 'test.xlsx'
		self.version = version
		self.default_manifest_path = 'manifest/{}/default.xml'.format(self.version)
		self.current_path = os.getcwd()
		self.excel_path = os.path.join(self.current_path, self.file_name)
		self.manifest_path = os.path.join(self.current_path, self.default_manifest_path)
		self.clone = 'git clone https://github.com/Bonny10/manifest.git'

	def is_excel_available(self):
		if (os.path.exists(self.excel_path)):
			msg = "Manifest excel file is available"
			print(msg)
			return True
		else:
			print("Please copy the Manifest excel in current directory")
			raw_input("Enter any key to continue..")
			sys.exit()

	def open_workspace(self):
		if(self.is_excel_available):
			self.workspace_obj = openpyxl.load_workbook(self.excel_path)
			if(self.workspace_obj):
				print("Opening Manifest excel workspace")
				return self.workspace_obj
			else:
				print("Not able to open Manifest excel workspace")
				raw_input("Enter any key to continue..")
				sys.exit()

	def open_sheet(self):
		self.workspace_obj=self.open_workspace()
		if (self.workspace_obj):
			self.sheet_obj = self.workspace_obj.active
			if(self.sheet_obj):
				print("Opening active sheet of Manifest Excel")
				return self.sheet_obj
			else:
				print("Not able to open active sheet of Manifest Excel")
				raw_input("Enter any key to continue..")
				sys.exit()

	def is_manifest_available(self):
		print(os.path.exists(self.manifest_path))
		if (os.path.exists(self.manifest_path)):
			print("Manifest is in current directory")
			print("Start modifying manifest.xml according to manifest.xlsx")
			self.parser = CommentedTreeBuilder()
			self.manifest_tree = ET.parse(self.manifest_path, self.parser)
			self.manifest_obj = self.manifest_tree.getroot()
		else:
			print("Manifest is not available cloning manifest")
			time.sleep(1)
			os.system(self.clone)
			self.parser = CommentedTreeBuilder()
			self.manifest_tree = ET.parse(self.manifest_path, self.parser)
			self.manifest_obj = self.manifest_tree.getroot()


	def update_manifest(self):
		self.sheet_obj=self.open_sheet()
		self.maxium_row_available = self.sheet_obj.max_row
		for i in range(2, self.maxium_row_available + 1):
			
			self.repo_path_obj = self.sheet_obj.cell(row = i, column = 1)
			self.repo_revision_obj = self.sheet_obj.cell(row = i, column = 2)
			self.repo_branch_obj = self.sheet_obj.cell(row = i, column = 3)
			
			self.repository_path = str(self.repo_path_obj.value)
			self.repository_revision = str(self.repo_revision_obj.value)
			self.repository_branch = str(self.repo_branch_obj.value)

			
			for self.manifest_child in self.manifest_obj:
				self.xml_manifest_path = self.manifest_child.get('path')
				if (self.repository_path == self.xml_manifest_path):
					print("\n Update manifest for : {}".format(self.repository_path))
					self.manifest_child.set('revision', self.repository_revision)
					self.manifest_child.set('upstream', self.repository_branch)
		
		self.manifest_tree.write('default_updated.xml', encoding="UTF-8")

	# def __del__(self):
	# 	self.manifest_tree.parse.close()

if __name__ == '__main__':
	version = raw_input('Enter version which manifest need to be changed: ')
	manifest = Manifest_update(version)
	manifest.is_excel_available()
	manifest.is_manifest_available()
	manifest.update_manifest()
